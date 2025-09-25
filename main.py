import os
import shutil
import datetime
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

warnings.filterwarnings('ignore')


def has_existing_data(folder: str) -> bool:
    if not os.path.isdir(folder):
        return False
    for name in os.listdir(folder):
        if not name.lower().endswith('.xlsx'):
            continue
        parts = os.path.splitext(name)[0].split('_')
        # 仅接受 “公司_财务_yyyymm.xlsx” 或 “公司_业务_yyyymm.xlsx”
        if len(parts) == 3 and parts[1] in ('财务', '业务'):
            return True
    return False


def backup_existing_data(data_folder: str, report_folder: str) -> str:
    """Copy all .xlsx files from data_folder to a timestamped backup directory under report_folder/data_backup.

    Returns the path to the created backup directory (or an empty string if nothing copied).
    """
    if not os.path.isdir(data_folder):
        return ""
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_root = os.path.join(report_folder, 'data_backup')
    backup_dir = os.path.join(backup_root, timestamp)
    os.makedirs(backup_dir, exist_ok=True)

    copied = 0
    for name in os.listdir(data_folder):
        if not name.lower().endswith('.xlsx'):
            continue
        src = os.path.join(data_folder, name)
        if os.path.isfile(src):
            dst = os.path.join(backup_dir, name)
            shutil.copy2(src, dst)
            copied += 1

    if copied == 0:
        # Remove empty timestamped folder to avoid clutter
        try:
            os.rmdir(backup_dir)
        except OSError:
            pass
        return ""
    return backup_dir


class PerformanceAnalyzer:
    def __init__(self, data_path, report_path, score_rules):
        self.data_path = data_path
        self.report_path = report_path
        self.score_rules = score_rules
        self.all_data = pd.DataFrame()
        self.kpi_data = pd.DataFrame()

        self.financial_columns = ['营业收入', '营业成本', '市场费用', '研发费用', '管理费用']
        self.business_columns = ['月初用户数', '月末用户数', '新增用户数', '总订单数', '付费用户数']

    # 1) 生成更真实模拟数据（若目录已有你的真实数据，程序会自动跳过）
    def create_realistic_dummy_data(self):
        if not os.path.exists(self.data_path):
            os.makedirs(self.data_path)

        rng = np.random.default_rng(20240925)

        companies = ['公司A', '公司B', '公司C']
        company_profiles = {
            '公司A': {
                'retention_base': 0.90, 'gross_margin_base': 0.62,
                'price_base': 160.0, 'paid_conv_base': 0.045, 'buys_per_paid': 2.3,
                'mkt_to_new_eff': 520.0, 'mkt_decay': 0.00008,
                'mkt_spend_base': 18000.0, 'mkt_spend_growth': 0.08
            },
            '公司B': {
                'retention_base': 0.82, 'gross_margin_base': 0.50,
                'price_base': 130.0, 'paid_conv_base': 0.035, 'buys_per_paid': 1.9,
                'mkt_to_new_eff': 380.0, 'mkt_decay': 0.00012,
                'mkt_spend_base': 20000.0, 'mkt_spend_growth': 0.10
            },
            '公司C': {
                'retention_base': 0.74, 'gross_margin_base': 0.42,
                'price_base': 110.0, 'paid_conv_base': 0.028, 'buys_per_paid': 1.5,
                'mkt_to_new_eff': 260.0, 'mkt_decay': 0.00018,
                'mkt_spend_base': 22000.0, 'mkt_spend_growth': 0.14
            }
        }

        def diminishing_returns(spend, a, b):
            spend = max(0.0, float(spend))
            return a * (1.0 - np.exp(-b * spend))

        for company in companies:
            p = company_profiles[company]

            month_start_users = 40000.0  # 7月初用户
            for month in range(7, 10):  # 2023-07, 08, 09
                date_str = f'2023{month:02}'
                step = month - 7  # 0,1,2

                mkt_spend = p['mkt_spend_base'] * ((1.0 + p['mkt_spend_growth']) ** step)
                mkt_spend *= np.clip(np.random.normal(1.0, 0.05), 0.85, 1.20)

                paid_new_expected = diminishing_returns(mkt_spend, a=p['mkt_to_new_eff'], b=p['mkt_decay'])
                paid_new = max(0, int(np.random.poisson(max(1.0, paid_new_expected))))
                organic_new = max(0, int(np.random.poisson(800 + 200 * step)))
                new_users = paid_new + organic_new

                retention = np.clip(p['retention_base'] + np.random.normal(0, 0.015), 0.65, 0.96)
                month_end_users = month_start_users * retention + new_users

                price = max(60.0, p['price_base'] * np.clip(np.random.normal(1.0, 0.04), 0.85, 1.20))
                paid_conv = np.clip(p['paid_conv_base'] + np.random.normal(0, 0.003) - 0.00002 * (price - 120.0),
                                    0.01, 0.10)
                paid_users = int(month_end_users * paid_conv)
                buys_per_paid = max(0.8, np.random.normal(p['buys_per_paid'], 0.15))

                revenue = paid_users * price * buys_per_paid
                gross_margin = np.clip(p['gross_margin_base'] + np.random.normal(0, 0.02), 0.30, 0.75)
                cost = revenue * (1.0 - gross_margin)

                rd_fee = max(8000.0, 12000.0 + step * 600.0 + np.random.normal(0, 600))
                admin_fee = max(5000.0, 6000.0 + step * 250.0 + np.random.normal(0, 300))

                total_orders = int(paid_users * buys_per_paid)

                financial_df = pd.DataFrame({
                    '科目': self.financial_columns,
                    '金额': [
                        revenue,        # 营业收入
                        cost,           # 营业成本
                        mkt_spend,      # 市场费用
                        rd_fee,         # 研发费用
                        admin_fee       # 管理费用
                    ]
                })
                financial_df.to_excel(os.path.join(self.data_path, f'{company}_财务_{date_str}.xlsx'),
                                      index=False)

                business_df = pd.DataFrame({
                    '指标': self.business_columns,
                    '数值': [
                        month_start_users,      # 月初用户数
                        month_end_users,        # 月末用户数
                        new_users,              # 新增用户数（投放+自然）
                        total_orders,           # 总订单数
                        paid_users              # 付费用户数
                    ]
                })
                business_df.to_excel(os.path.join(self.data_path, f'{company}_业务_{date_str}.xlsx'),
                                     index=False)

                month_start_users = month_end_users

    # 2) 读取与整合（去重、统一列、强制类型）
    def load_and_integrate_data(self):
        financial_rows = []
        business_rows = []

        try:
            if not os.path.isdir(self.data_path):
                print(f'数据目录不存在: {self.data_path}')
                return False

            for filename in os.listdir(self.data_path):
                if not filename.lower().endswith('.xlsx'):
                    continue
                parts = os.path.splitext(filename)[0].split('_')
                if len(parts) != 3:
                    continue

                company, data_type, date_str = parts
                file_path = os.path.join(self.data_path, filename)
                df = pd.read_excel(file_path, engine='openpyxl')

                if len(df.columns) != 2:
                    continue

                key_col, value_col = df.columns[0], df.columns[1]
                df[key_col] = df[key_col].astype(str).str.strip()
                df[value_col] = pd.to_numeric(df[value_col], errors='coerce')

                if data_type == '财务':
                    row = {'公司': company, '日期': date_str}
                    for _, r in df.iterrows():
                        if r[key_col] in self.financial_columns:
                            row[r[key_col]] = r[value_col]
                    financial_rows.append(row)
                elif data_type == '业务':
                    row = {'公司': company, '日期': date_str}
                    for _, r in df.iterrows():
                        if r[key_col] in self.business_columns:
                            row[r[key_col]] = r[value_col]
                    business_rows.append(row)

            financial_df = pd.DataFrame(financial_rows)
            business_df = pd.DataFrame(business_rows)

            for col in ['公司', '日期'] + self.financial_columns:
                if col not in financial_df.columns:
                    financial_df[col] = np.nan
            for col in ['公司', '日期'] + self.business_columns:
                if col not in business_df.columns:
                    business_df[col] = np.nan

            if not financial_df.empty:
                financial_df = (financial_df.sort_values(['公司', '日期'])
                                .drop_duplicates(['公司', '日期'], keep='last'))
            if not business_df.empty:
                business_df = (business_df.sort_values(['公司', '日期'])
                               .drop_duplicates(['公司', '日期'], keep='last'))

            merged = pd.merge(financial_df, business_df, on=['公司', '日期'], how='outer', suffixes=('', '_dup'))

            for col in self.business_columns:
                dup = f'{col}_dup'
                if dup in merged.columns:
                    merged[col] = merged[col].fillna(merged[dup])
                    merged.drop(columns=[dup], inplace=True, errors='ignore')

            for col in self.financial_columns + self.business_columns:
                if col in merged.columns:
                    merged[col] = pd.to_numeric(merged[col], errors='coerce')

            cols_order = ['公司', '日期'] + self.financial_columns + self.business_columns
            for c in cols_order:
                if c not in merged.columns:
                    merged[c] = np.nan

            self.all_data = merged[cols_order].copy()
            return True
        except Exception as e:
            print(f'数据整合错误: {e}')
            import traceback
            traceback.print_exc()
            return False

    # 3) KPI 计算（解耦口径 + 宽护栏）
    def calculate_realistic_kpis(self):
        if self.all_data.empty:
            print('错误: 没有数据可供计算')
            return False
        try:
            kpi = self.all_data.copy()
            kpi['日期'] = pd.to_datetime(kpi['日期'], format='%Y%m', errors='coerce')
            kpi = kpi.dropna(subset=['日期']).sort_values(['公司', '日期'])

            for col in self.financial_columns + self.business_columns:
                kpi[col] = pd.to_numeric(kpi[col], errors='coerce')

            kpi['上月营收'] = kpi.groupby('公司')['营业收入'].shift(1)
            kpi['营收月度增长率'] = np.where(
                (kpi['上月营收'] > 0) & kpi['上月营收'].notna(),
                (kpi['营业收入'] - kpi['上月营收']) / kpi['上月营收'],
                np.nan
            )

            # CAC：仅用投放新增估算（市场费用 / 估计的投放新增）
            est_organic = np.clip(kpi['月初用户数'] * 0.01, 200, 3000)
            est_paid_new = np.clip(kpi['新增用户数'] - est_organic, 1.0, None)
            kpi['获客成本(CAC)'] = np.where(est_paid_new > 0, kpi['市场费用'] / est_paid_new, np.nan)

            kpi['上月月末用户数'] = kpi.groupby('公司')['月末用户数'].shift(1)
            kpi['用户留存率'] = np.where(
                (kpi['上月月末用户数'] > 0) & kpi['上月月末用户数'].notna(),
                (kpi['月末用户数'] - kpi['新增用户数']) / kpi['上月月末用户数'],
                np.nan
            )
            kpi['用户留存率'] = np.clip(kpi['用户留存率'], 0.55, 0.98)

            kpi['毛利率'] = np.where(kpi['营业收入'] > 0,
                                   (kpi['营业收入'] - kpi['营业成本']) / kpi['营业收入'],
                                   np.nan)
            kpi['毛利率'] = np.clip(kpi['毛利率'], 0.20, 0.85)

            kpi['ARPU'] = np.where(kpi['付费用户数'] > 0, kpi['营业收入'] / kpi['付费用户数'], np.nan)

            kpi['LTV（估算）'] = np.where(
                kpi[['ARPU', '毛利率', '用户留存率']].notna().all(axis=1),
                kpi['ARPU'] * kpi['毛利率'] * (1.0 / np.maximum(0.05, (1.1 - kpi['用户留存率']))),
                np.nan
            )
            kpi['LTV（估算）'] = np.clip(kpi['LTV（估算）'], 30, 100000)

            kpi['LTV/CAC 比率'] = np.where(
                (kpi['LTV（估算）'].notna()) & (kpi['获客成本(CAC)'] > 0),
                kpi['LTV（估算）'] / kpi['获客成本(CAC)'],
                np.nan
            )
            kpi['LTV/CAC 比率'] = np.clip(kpi['LTV/CAC 比率'], 0.01, 100.0)

            self.kpi_data = kpi
            return True
        except Exception as e:
            print(f'KPI计算错误: {e}')
            import traceback
            traceback.print_exc()
            return False

    # 4) 健康度评分
    def calculate_scores(self):
        if self.kpi_data.empty:
            print('错误: KPI数据为空')
            return False
        try:
            kpi = self.kpi_data.reset_index(drop=True).copy()

            def row_score(row):
                total = 0
                gr = row.get('营收月度增长率', np.nan)
                if pd.notna(gr):
                    if gr > self.score_rules['营收月度增长率']['高']:
                        total += 5
                    elif gr > self.score_rules['营收月度增长率']['中']:
                        total += 3
                    else:
                        total += 1

                ratio = row.get('LTV/CAC 比率', np.nan)
                if pd.notna(ratio):
                    if ratio > self.score_rules['LTV/CAC 比率']['高']:
                        total += 5
                    elif ratio > self.score_rules['LTV/CAC 比率']['中']:
                        total += 3
                    else:
                        total += 1

                if total >= 8:
                    rating = 'A[优秀]'
                elif total >= 5:
                    rating = 'B[良好]'
                elif total >= 3:
                    rating = 'C[及格]'
                else:
                    rating = 'D[危险]'
                return total, rating

            scores, ratings = [], []
            for _, r in kpi.iterrows():
                s, rt = row_score(r)
                scores.append(s)
                ratings.append(rt)

            kpi['综合得分'] = scores
            kpi['健康度评级'] = ratings

            self.kpi_data = kpi
            return True
        except Exception as e:
            print(f'评分计算错误: {e}')
            import traceback
            traceback.print_exc()
            return False

    # 5) 生成报告（仪表盘用数值+格式；图表引用数值；系列标题用 SeriesLabel）
    def generate_enhanced_report(self):
        if self.kpi_data.empty:
            print('KPI数据为空，无法生成报告。')
            return False

        if not os.path.exists(self.report_path):
            os.makedirs(self.report_path)
        today = datetime.date.today().strftime('%Y%m%d')
        report_filename = os.path.join(self.report_path, f'分公司效能分析报告_{today}.xlsx')

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = '汇总仪表盘'

            ws['A1'] = '分公司业务效能分析报告'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            latest = (self.kpi_data.sort_values('日期')
                      .groupby('公司', as_index=False)
                      .last())
            latest = latest.copy()
            latest['日期文本'] = latest['日期'].dt.strftime('%Y-%m')

            headers = ['公司', '日期文本', '营业收入', '营收月度增长率', 'LTV/CAC 比率', '综合得分', '健康度评级']
            ws['A4'] = '各分公司最新KPI数据'
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=i, value=h)
                cell.font = Font(bold=True)

            start_row = 6
            for r_idx, (_, row) in enumerate(latest.iterrows(), start=start_row):
                ws.cell(row=r_idx, column=1, value=row['公司'])
                ws.cell(row=r_idx, column=2, value=row['日期文本'])

                c_rev = ws.cell(row=r_idx, column=3, value=self._to_float(row.get('营业收入')))
                c_rev.number_format = '#,##0'

                c_gr = ws.cell(row=r_idx, column=4, value=self._to_float(row.get('营收月度增长率')))
                c_gr.number_format = '0.00%'

                c_ratio = ws.cell(row=r_idx, column=5, value=self._to_float(row.get('LTV/CAC 比率')))
                c_ratio.number_format = '0.00'

                ws.cell(row=r_idx, column=6, value=self._to_float(row.get('综合得分')))
                ws.cell(row=r_idx, column=7, value=row.get('健康度评级'))

            n = len(latest)
            if n > 0:
                first = start_row
                last = start_row + n - 1

                chart1 = BarChart()
                chart1.title = '各分公司LTV/CAC比率对比'
                chart1.x_axis.title = '分公司'
                chart1.y_axis.title = '比率'
                data_ref = Reference(ws, min_col=5, max_col=5, min_row=first, max_row=last)
                cats_ref = Reference(ws, min_col=1, max_col=1, min_row=first, max_row=last)
                chart1.add_data(data_ref, titles_from_data=False)
                chart1.set_categories(cats_ref)
                if chart1.series:
                    chart1.series[0].title = SeriesLabel(v='LTV/CAC 比率')
                ws.add_chart(chart1, 'H5')

                chart2 = BarChart()
                chart2.title = '各分公司营业收入对比'
                chart2.x_axis.title = '分公司'
                chart2.y_axis.title = '收入(元)'
                data_ref2 = Reference(ws, min_col=3, max_col=3, min_row=first, max_row=last)
                cats_ref2 = Reference(ws, min_col=1, max_col=1, min_row=first, max_row=last)
                chart2.add_data(data_ref2, titles_from_data=False)
                chart2.set_categories(cats_ref2)
                if chart2.series:
                    chart2.series[0].title = SeriesLabel(v='营业收入')
                ws.add_chart(chart2, 'H20')

            ws_d = wb.create_sheet('明细数据')
            detail = self.kpi_data.copy()
            detail['日期'] = detail['日期'].dt.strftime('%Y-%m')
            for r in dataframe_to_rows(detail, index=False, header=True):
                ws_d.append(r)

            for col in ws_d.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    v = '' if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws_d.column_dimensions[col_letter].width = min(max_len + 2, 50)

            wb.save(report_filename)
            print(f'报告生成完毕: {report_filename}')
            return True
        except Exception as e:
            print(f'报告生成错误: {e}')
            import traceback
            traceback.print_exc()
            return False

    @staticmethod
    def _to_float(v, default=None):
        if v is None:
            return default
        try:
            if isinstance(v, float) and np.isnan(v):
                return default
            return float(v)
        except Exception:
            return default


def main():
    config = {
        'data_folder': r'D:\分公司月度数据',
        'report_folder': r'D:\效能分析报告',
        'score_rules': {
            '营收月度增长率': {'高': 0.15, '中': 0.05},
            'LTV/CAC 比率': {'高': 3, '中': 1},
        }
    }

    analyzer = PerformanceAnalyzer(
        data_path=config['data_folder'],
        report_path=config['report_folder'],
        score_rules=config['score_rules']
    )

    print('=== 步骤1: 数据准备 ===')
    if has_existing_data(config['data_folder']):
        print(f"检测到 {config['data_folder']} 已有数据文件，先创建时间戳副本再分析。")
        backup_dir = backup_existing_data(config['data_folder'], config['report_folder'])
        if backup_dir:
            print(f"已备份数据到: {backup_dir}")
        else:
            print("未发现可备份的 .xlsx 文件或备份目录为空。")
    else:
        print("未检测到数据文件，先生成模拟数据。")
        analyzer.create_realistic_dummy_data()

    print('\n=== 步骤2: 数据整合 ===')
    if not analyzer.load_and_integrate_data():
        return

    print('\n=== 步骤3: KPI计算 ===')
    if not analyzer.calculate_realistic_kpis():
        return

    print('\n=== 步骤4: 健康度评分 ===')
    if not analyzer.calculate_scores():
        return

    print('\n=== 步骤5: 生成报告 ===')
    analyzer.generate_enhanced_report()

    print('\n=== 程序执行完毕 ===')


if __name__ == '__main__':
    main()
